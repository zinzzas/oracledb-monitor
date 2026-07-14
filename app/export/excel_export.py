"""
엑셀(.xlsx) 다운로드 생성 모듈.

바인드 파라미터값 관련 중요한 한계:
V$SQL_BIND_CAPTURE는 커서당 최대 15분 간격으로만 값을 샘플링하고, WHERE/HAVING절에
쓰인 단순 타입 바인드만 캡처한다 (LONG/LOB/객체/REF CURSOR 제외). 즉 "이 특정 실행"의
정확한 값이 아니라 "최근 어느 시점에 캡처된 값"이며, 이 프로젝트가 다뤄온 MyBatis
REF CURSOR OUT 패턴은 애초에 캡처 대상이 아니다. 그래서 참고용(최선노력) 컬럼으로만
제공하고, 시트 안내문에도 명시한다.
"""
from __future__ import annotations
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

TIER_LABEL = {"lt10": "3~10초", "lt15": "10~15초", "ge15": "15초 이상"}


def _write_sheet(ws, headers: list[str], rows: list[list], widths: list[int] | None = None):
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    if widths:
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"


def build_long_session_excel(rows: list[dict], start_ts: int, end_ts: int) -> bytes:
    """Long Active Session Count 드릴다운 결과를 엑셀로. rows 의 각 dict 는
    sqlite_store.get_long_query_log() 반환 형태(+선택적으로 bind_params_text)를 기대한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Long Active Sessions"

    headers = ["실행시각", "User", "Module", "구간(Tier)", "Elapsed(초)", "PL/SQL 여부",
               "SQL_ID", "SQL Text / 패키지.프로시저", "Bind Params (참고용, 부정확할 수 있음)"]
    data_rows = []
    for r in rows:
        ts_label = datetime.fromtimestamp(r["ts"]).strftime("%Y-%m-%d %H:%M:%S")
        data_rows.append([
            ts_label,
            r.get("username") or "",
            r.get("module") or "",
            TIER_LABEL.get(r.get("tier"), r.get("tier") or ""),
            r.get("elapsed_sec"),
            "O" if r.get("is_plsql_call") else "",
            r.get("sql_id") or "",
            r.get("sql_text") or "",
            r.get("bind_params_text") or "",
        ])
    _write_sheet(ws, headers, data_rows, widths=[19, 14, 16, 12, 12, 10, 16, 60, 40])

    note = wb.create_sheet("안내")
    note["A1"] = "조회 구간"
    note["B1"] = (
        f"{datetime.fromtimestamp(start_ts).strftime('%Y-%m-%d %H:%M:%S')} ~ "
        f"{datetime.fromtimestamp(end_ts).strftime('%Y-%m-%d %H:%M:%S')}"
    )
    note["A2"] = "Bind Params 안내"
    note["B2"] = (
        "V$SQL_BIND_CAPTURE 기반 최선노력 값입니다. Oracle이 커서당 최대 15분 간격으로만 샘플링하고, "
        "REF CURSOR/OUT 파라미터·LOB 등은 애초에 캡처되지 않습니다. 이 실행의 정확한 값이 아닐 수 있습니다."
    )
    note.column_dimensions["A"].width = 18
    note.column_dimensions["B"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_plsql_calls_excel(rows: list[dict], start_ts: int, end_ts: int) -> bytes:
    """PL/SQL Package Calls 상세 조회 결과를 엑셀로."""
    wb = Workbook()
    ws = wb.active
    ws.title = "PL-SQL Package Calls"

    headers = ["실행시각", "Package.Procedure", "Calls", "Avg(ms)", "Total(sec)"]
    data_rows = []
    for r in rows:
        ts_label = datetime.fromtimestamp(r["ts"]).strftime("%Y-%m-%d %H:%M:%S")
        data_rows.append([
            ts_label,
            r.get("proc_name") or "",
            r.get("calls"),
            r.get("avg_elapsed_ms"),
            round((r.get("total_elapsed_ms") or 0) / 1000, 2),
        ])
    _write_sheet(ws, headers, data_rows, widths=[19, 50, 10, 12, 12])

    note = wb.create_sheet("안내")
    note["A1"] = "조회 구간"
    note["B1"] = (
        f"{datetime.fromtimestamp(start_ts).strftime('%Y-%m-%d %H:%M:%S')} ~ "
        f"{datetime.fromtimestamp(end_ts).strftime('%Y-%m-%d %H:%M:%S')}"
    )
    note["A2"] = "집계 방식"
    note["B2"] = "V$SQL 누적 카운터(EXECUTIONS/ELAPSED_TIME) 델타 기반 - 폴링 손실 없이 정확히 집계됩니다."
    note.column_dimensions["A"].width = 18
    note.column_dimensions["B"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
